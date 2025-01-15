import os 
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from pypdf import PdfReader

clients = set()

def restore_clients(file_path):
    global clients
    path = os.path.join(file_path, "clients_2023.txt")
    if os.path.exists(path):
        with open(path, "r") as f:
            clients = set(f.read().split("\n---\n"))
def save_clients(file_path):
    global clients
    path = os.path.join(file_path, "clients_2023.txt")
    with open(path, "a") as f:
        f.write("\n---\n".join(clients))
        f.write("\n---\n")
def extract_invoice_data(file_path):
    
    with PdfReader(file_path) as reader:
        print(f"attempting to read {file_path}")
        try:
            fields = reader.get_form_text_fields()
        except:
            print(f"extract_invoice_data: Error reading {file_path}")
            return ""
        if 'Bill To' not in fields:
            print(f"extract_invoice_data: No 'Bill To' field found in {file_path}")
            return ""
        bill_to_field = fields['Bill To']
        if not bill_to_field:
            print(f"extract_invoice_data: No 'Bill To' field found in {file_path}")
            return ""
        newline_bill_to = bill_to_field.replace("\r", "\n")
        print(f"this is the extracted bill to field with carriages{bill_to_field} \n then replaced: {newline_bill_to}")
        
        
    return newline_bill_to
def format_client_data(text):
    bill_lines = text.split("\n")
    lines = []
    for line in bill_lines:
        phone_number_match = re.search(r'\(?(\d{3})\)?(-|\s|\.)?(\d{3})(-|\s|\.)?(\d{4})', line)
        email_match = re.search(r'[\w\+\.%-]+@[\w\.-]+', line)
        if phone_number_match:
            print(f"this is the phone number match {phone_number_match.group(0)}")
            continue
        if email_match:
            print(f"this is the email match {email_match.group(0)}")
            continue
        lines.append(line)
            
    if len(lines) < 5:
        lines.insert(0, "\n")
    new_text = "\n".join(lines)
    print(f"this is the formatted text {new_text}")
    if new_text in clients:
        print("format_client_data: Data already in clients.txt")
        return ""
    else:
        clients.add(new_text)
    return new_text

def find_invoice_pdfs(base_path,test = False):
    """
    Find all invoice PDFs in the given directory, starting from the most recent truck folder.
    """
    # folders = [f for f in os.listdir(base_path) if os.path.isdir(f)]
    folders = os.listdir(base_path)
    print(folders, os.listdir(base_path))
    numbered_folders = []
    for f in folders:
        match = re.search(r'(\d+)$', f)
        if match:
            print(f"find_invoice_pdfs: Found truck folder: {f}")
            number = int(match.group(1))
            numbered_folders.append((number, f))
    
    numbered_folders.sort(reverse=True)
    invoice_files = []
    for i in range(len(numbered_folders)):
        folder = numbered_folders[i][1]
        print(f"find_invoice_pdfs: Checking folder: {folder}")
        for root, dirs, files in os.walk(os.path.join(base_path, folder)):
            for file in files:
                match = re.search(r'.*(Inv|Invoice).*\.pdf$', file, re.IGNORECASE) 
                
                if match:
                    print(f"find_invoice_pdfs: Found invoice PDF: {file}")
                    invoice_files.append(os.path.join(root, file))
                    if test:
                        break
            if test and len(invoice_files) > 0:
                break
        print(f"find_invoice_pdfs: finished checking folder: {folder}")
    return invoice_files
def create_client_data(invoice_files, base_path, output_path, name, test=False):
    """
    Extract client data from each invoice PDF and store it in a spreadsheet.
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = str(re.search(r'(\d+)$', base_path).group(1)) + "_Client_Address_Data"
    row = 1
    sheet.cell(row=row, column=1, value = "Client Address")
    row += 1
    for file in invoice_files:
        text = extract_invoice_data(file)
        print(f"create_client_data: Extracted client data from {file}")
        if not text:
            print(f"create_client_data: No client data found in {file}")
            continue
        data = format_client_data(text)
        if not data:
            print(f"create_client_data: Duplicate client data found in {file}")
            continue
        cell = sheet.cell(row=row, column=1, value = data)
        cell.alignment = Alignment(wrap_text=True)
        
        row += 1
    
    wb.save(os.path.join(output_path, name))
if __name__ == '__main__':
    
    # Set up command-line argument parsing
    parser = argparse.ArgumentParser(description="Download and organize client data from invoices by date.")
    parser.add_argument("--root", required=True, help="The root directory where invoices are stored.")
    parser.add_argument("--output", required=True, help="The output directory where the client data will be stored.")
    parser.add_argument("--test", required=False, help="Whether to run in test mode.", action="store_true")
    parser.add_argument("--name", required=False, help="What you want to name the excel file output. MUST end in .xlsx")
    
    args = parser.parse_args()
    
    print(f"{'main: Simulating first client extract' if args.test else 'main: Extracting client data from invoices'} to: {args.output}")
    
    # parser.add_argument("--file", required=True, help="The path to the invoice file.")
    # args = parser.parse_args()
    # text = extract_invoice_data(args.file)
    # print(text)
    
    # run the main code
    restore_clients(args.output)
    invoice_files = find_invoice_pdfs(args.root, args.test)
    create_client_data(invoice_files, args.root, args.output, args.name, args.test)
    save_clients(args.output)
    print("Done!")


